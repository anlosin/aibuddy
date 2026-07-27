"""Excel 处理插件 — 创建/读取电子表格"""
import os
import json

PLUGIN_INFO = {
    "name": "excel",
    "description": "创建 Excel 表格文件或读取已有表格数据。创建时需要提供文件名、表头和行数据；读取时需要提供文件路径。",
    "version": "1.0",
}

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "create_excel",
            "description": "创建一个 Excel 文件（.xlsx）。需要提供文件名、列标题和数据行。文件保存在程序同级目录下。",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "文件名，如 report.xlsx（不要包含路径，只写文件名）"
                    },
                    "headers": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "列表头，如 ['姓名', '年龄', '城市']"
                    },
                    "rows": {
                        "type": "array",
                        "items": {"type": "array", "items": {"type": "string"}},
                        "description": "数据行，每行是一个数组，如 [['张三','25','北京'], ['李四','30','上海']]"
                    }
                },
                "required": ["filename", "headers", "rows"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_excel",
            "description": "读取一个 Excel 文件的内容，返回表头和前100行数据。文件路径相对于程序目录或绝对路径。",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径，如 report.xlsx 或绝对路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "要读取的工作表名称（可选，默认读取第一个工作表）"
                    }
                },
                "required": ["filepath"]
            }
        }
    },
]


def _safe_path(filepath):
    """解析文件路径"""
    if os.path.isabs(filepath):
        return filepath
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", filepath)


def _format_table(headers, rows):
    """将表格数据格式化为可读文本"""
    lines = []
    # 表头
    lines.append(" | ".join(str(h) for h in headers))
    lines.append("-" * 50)
    # 数据行
    for row in rows:
        lines.append(" | ".join(str(c) for c in row))
    return "\n".join(lines)


def execute(name, arguments):
    if name == "create_excel":
        return _create(arguments)
    if name == "read_excel":
        return _read(arguments)
    return f"未知工具: {name}"


def _create(args):
    try:
        import openpyxl
    except ImportError:
        return "错误: 请先安装 openpyxl (pip install openpyxl)"

    filename = args.get("filename", "output.xlsx")
    headers = args.get("headers", [])
    rows = args.get("rows", [])

    if not headers:
        return "错误: 表头不能为空"

    path = _safe_path(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        from openpyxl.styles import Font, PatternFill
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 写数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                w = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, w)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(path)
    return (
        f"Excel 文件已创建: {filename}\n"
        f"路径: {os.path.abspath(path)}\n"
        f"表: Sheet1, {len(rows)} 行 × {len(headers)} 列"
    )


def _read(args):
    try:
        import openpyxl
    except ImportError:
        return "错误: 请先安装 openpyxl (pip install openpyxl)"

    filepath = args.get("filepath", "")
    sheet_name = args.get("sheet_name")

    if not filepath:
        return "错误: 未提供文件路径"

    path = _safe_path(filepath)
    if not os.path.exists(path):
        return f"错误: 文件不存在 - {path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return f"无法打开文件: {e}"

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return f"工作表 '{sheet_name}' 不存在，可用工作表: {', '.join(wb.sheetnames)}"
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    # 读取数据（最多100行）
    headers = []
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=101, values_only=True)):
        if row_idx == 0:
            headers = [str(c) if c is not None else f"列{i+1}" for i, c in enumerate(row)]
        else:
            rows.append([str(c) if c is not None else "" for c in row])

    table = _format_table(headers, rows)
    return (
        f"[{filepath}] 工作表: {sheet_name}\n"
        f"总行数: {ws.max_row}, 总列数: {ws.max_column}\n"
        f"\n{table}"
    )
